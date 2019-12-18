"""
The template of the script for the machine learning process in game pingpong
"""

# Import the necessary modules and classes
import pygame
from mlgame import gameconfig
import importlib
import games.pingpong.communication as comm
from games.pingpong.communication import (
    SceneInfo, GameInstruction, GameStatus, PlatformAction
)
import pickle
import numpy as np
from openpyxl import load_workbook

data = []
loadFile = open("unique_data_P1.xlsx", "rb")
wb = load_workbook(loadFile)
sheet = wb.active 
rows = list(sheet.rows)

for i in range(1, len(rows)):
    buffer = []
    for j in range(5):
        buffer.append(rows[i][j].value)
    data.append(buffer)
loadFile.close()
data = np.array(data)

from os import listdir
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws['A1'] = 'start_x'
ws['B1'] = 'speed'
ws['C1'] = 'm'
ws['D1'] = 'frame_remainder'
ws['E1'] = 'end_x'

def predict(parameter):
    parameter_length = len(parameter)
    y = [ x for x in data if np.all(x[:parameter_length] == parameter[:parameter_length])]
    aid_length = len(y)
    if aid_length != 0:
        aid = 0
        for aid_value in y:
            aid += aid_value[-1]
        aid = round(aid / len(y))
        return aid
    else:
        return predict(parameter[:-1])

def ml_loop(side: str):
    comm.ml_ready()
    past_ball_position = []
    ball_down = False
    first = True
    aid = 100
    while True:
        scene_info = comm.get_scene_info()
        if scene_info.status == GameStatus.GAME_1P_WIN or \
           scene_info.status == GameStatus.GAME_2P_WIN:
            comm.ml_ready()
            first = True
            dist = "new_log1"
            file_length = len(listdir(dist))
            wb.save(dist + '/7th_' + str(file_length + 1) + '.xlsx')
            continue

        now_ball_position = scene_info.ball
            
        if len(past_ball_position) == 0:
            past_ball_position = now_ball_position
            continue
        else:
            if (now_ball_position[1] - past_ball_position[1]) > 0:
                ball_down = True
            else:
                ball_down = False

        m = 0
        if now_ball_position[0] - past_ball_position[0] != 0:
            m = (now_ball_position[1] - past_ball_position[1]) / (now_ball_position[0] - past_ball_position[0])
        
        if ball_down == False:
            aid = 100

        if first:
            aid = round(now_ball_position[0] - ((now_ball_position[1] - 415) / m))
            if aid < 0:
                aid = -aid
            elif aid > 195:
                aid = 200 - (aid - 200)
        
        if ball_down == True:
            if now_ball_position[1] == 415:
                params.append(scene_info.ball[0])
                ws.append(params)
        else:
            if now_ball_position[1] == 80:
                params = [int(scene_info.ball[0]), int(scene_info.ball_speed), m, int(scene_info.frame % 200)]

        if now_ball_position[1] == 80:
            first = False
            parameter = [int(scene_info.ball[0]), round(m, 2), int(scene_info.ball_speed), int(scene_info.frame % 200)]
            aid = predict(parameter)
            cc = aid % 5
            aid -= cc

        now_platform_positionX = scene_info.platform_1P[0] + 20
        if aid > now_platform_positionX:
            comm.send_instruction(scene_info.frame, PlatformAction.MOVE_RIGHT)
        elif aid < now_platform_positionX:
            comm.send_instruction(scene_info.frame, PlatformAction.MOVE_LEFT)
        elif aid == now_platform_positionX:
            comm.send_instruction(scene_info.frame, PlatformAction.NONE)

        past_ball_position = now_ball_position
